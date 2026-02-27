import re
import unicodedata
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text


def _canonical_tokens(value: object) -> List[str]:
    normalized = _normalize_text(value)
    cleaned = re.sub(r"[^a-z0-9]+", " ", normalized)
    return [t for t in cleaned.split() if t]


def _find_brand_in_part(part: str, known_brands: Dict[Tuple[str, ...], str]) -> Optional[str]:
    part_tokens = _canonical_tokens(part)
    if not part_tokens:
        return None

    best_match = None
    best_len = 0
    part_text = " ".join(part_tokens)

    for brand_tokens, brand_key in known_brands.items():
        candidate = " ".join(brand_tokens)
        if candidate and re.search(rf"\b{re.escape(candidate)}\b", part_text):
            if len(brand_tokens) > best_len:
                best_match = brand_key
                best_len = len(brand_tokens)

    return best_match


def _remove_brand_from_text(value: str, brand: str) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""

    brand_tokens = _canonical_tokens(brand)
    if not brand_tokens:
        return raw

    pattern = r"\b" + r"\s+".join(re.escape(t) for t in brand_tokens) + r"\b"
    normalized = _normalize_text(raw)
    without_brand = re.sub(pattern, " ", normalized, count=1)
    without_brand = re.sub(r"\s+", " ", without_brand).strip()

    if without_brand:
        return without_brand
    return raw


def parse_brand_model(value: object, command_brands: List[str]) -> Tuple[str, str]:
    raw = str(value or "").strip()
    if not raw:
        return "", ""

    known_brands: Dict[Tuple[str, ...], str] = {}
    for brand in command_brands:
        tokens = tuple(_canonical_tokens(brand))
        if tokens:
            known_brands[tokens] = brand

    # Prefer explicit separators (common in files: "VT08F - Protrack").
    parts = [p.strip() for p in re.split(r"\s+-\s+", raw) if p.strip()]
    if len(parts) == 2:
        left_brand = _find_brand_in_part(parts[0], known_brands)
        right_brand = _find_brand_in_part(parts[1], known_brands)

        if left_brand and not right_brand:
            return left_brand, parts[1].lower()
        if right_brand and not left_brand:
            return right_brand, parts[0].lower()

    first_word = _canonical_tokens(raw[: raw.find(" ")] if " " in raw else raw)
    last_word = _canonical_tokens(raw.split(" ")[-1])

    if first_word:
        first_brand = known_brands.get(tuple(first_word))
        if first_brand:
            model = raw[len(raw.split(" ", 1)[0]) :].strip(" -_/,")
            return first_brand, model.lower()

    if last_word:
        last_brand = known_brands.get(tuple(last_word))
        if last_brand:
            model = raw[: -len(raw.split(" ")[-1])].strip(" -_/,")
            return last_brand, model.lower()

    brand = _find_brand_in_part(raw, known_brands)
    if brand:
        return brand, _remove_brand_from_text(raw, brand).lower()

    # Fallback: assume "modelo marca" if there are at least two tokens.
    tokens = raw.split()
    if len(tokens) >= 2:
        return tokens[-1].lower(), " ".join(tokens[:-1]).lower()

    return "", raw.lower()


def _find_header_columns(ws) -> Optional[Dict[str, Optional[int]]]:
    # Scan top rows to tolerate dashboard blocks and decorative headers.
    max_scan_rows = min(40, ws.max_row or 0)
    max_scan_cols = min(30, ws.max_column or 0)

    for row_idx in range(1, max_scan_rows + 1):
        mapping: Dict[str, Optional[int]] = {
            "header_row": row_idx,
            "phone_col": None,
            "status_col": None,
            "brand_model_col": None,
            "brand_col": None,
            "model_col": None,
            "error_col": None,
        }

        for col_idx in range(1, max_scan_cols + 1):
            header = _normalize_text(ws.cell(row=row_idx, column=col_idx).value)

            if header in {"telefono", "telefono movil", "telefono movil / sim", "telefono/sim", "sim"}:
                mapping["phone_col"] = col_idx
            elif header in {"status", "estado"}:
                mapping["status_col"] = col_idx
            elif header in {"marca/modelo", "marca modelo", "marcamodelo"}:
                mapping["brand_model_col"] = col_idx
            elif header == "marca":
                mapping["brand_col"] = col_idx
            elif header == "modelo":
                mapping["model_col"] = col_idx
            elif header == "error":
                mapping["error_col"] = col_idx

        has_phone = mapping["phone_col"] is not None
        has_brand_model = mapping["brand_model_col"] is not None or (
            mapping["brand_col"] is not None and mapping["model_col"] is not None
        )

        if has_phone and has_brand_model:
            return mapping

    return None


def load_devices(path: str, commands_config: Optional[Dict[str, object]] = None) -> pd.DataFrame:
    workbook = load_workbook(path)
    command_brands = list((commands_config or {}).keys())
    rows: List[Dict[str, object]] = []

    for sheet in workbook.worksheets:
        header = _find_header_columns(sheet)
        if not header:
            continue

        row_start = int(header["header_row"]) + 1
        for row_idx in range(row_start, sheet.max_row + 1):
            phone_raw = sheet.cell(row=row_idx, column=header["phone_col"]).value
            phone = str(phone_raw or "").strip()

            if header["brand_model_col"] is not None:
                brand_model_raw = sheet.cell(row=row_idx, column=header["brand_model_col"]).value
                brand, model = parse_brand_model(brand_model_raw, command_brands)
            else:
                brand_raw = sheet.cell(row=row_idx, column=header["brand_col"]).value
                model_raw = sheet.cell(row=row_idx, column=header["model_col"]).value
                brand = str(brand_raw or "").strip().lower()
                model = str(model_raw or "").strip().lower()

            if not phone and not brand and not model:
                continue

            current_status = ""
            if header["status_col"] is not None:
                current_status = str(sheet.cell(row=row_idx, column=header["status_col"]).value or "").strip()

            current_error = ""
            if header["error_col"] is not None:
                current_error = str(sheet.cell(row=row_idx, column=header["error_col"]).value or "").strip()

            rows.append(
                {
                    "Telefono": phone,
                    "Marca": brand,
                    "Modelo": model,
                    "Status": current_status,
                    "Error": current_error,
                    "__sheet": sheet.title,
                    "__row": row_idx,
                    "__status_col": header["status_col"],
                    "__error_col": header["error_col"],
                }
            )

    return pd.DataFrame(rows)


def save_devices(df: pd.DataFrame, path: str):
    workbook = load_workbook(path)

    for _, row in df.iterrows():
        sheet_name = row.get("__sheet")
        if not sheet_name or sheet_name not in workbook.sheetnames:
            continue

        excel_row = row.get("__row")
        if pd.isna(excel_row):
            continue

        ws = workbook[sheet_name]
        row_idx = int(excel_row)

        status_col = row.get("__status_col")
        if status_col and not pd.isna(status_col):
            ws.cell(row=row_idx, column=int(status_col)).value = row.get("Status", "")

        error_col = row.get("__error_col")
        if error_col and not pd.isna(error_col):
            ws.cell(row=row_idx, column=int(error_col)).value = row.get("Error", "")

    workbook.save(path)
