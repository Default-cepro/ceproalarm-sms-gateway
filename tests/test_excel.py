import openpyxl
import pytest

from src.storage.excel import load_devices, parse_brand_model, save_devices

# Commands config mirroring the real config/commands.json brands.
COMMANDS = {
    "wanwaytech": {"gs-900": {}},
    "protrack": {"vt08f": {}},
    "coban": {"403b": {}},
    "concox": {"gt06e": {}},
    "jimi": {"jm-vl01": {}},
}

HEADERS = ["Telefono", "Marca/Modelo", "Placas", "Status", "Error"]


def _make_workbook(path):
    """Build a workbook with decorative rows above the header and two sheets."""
    wb = openpyxl.Workbook()

    for sheet_name in ("Aragua", "Carabobo"):
        ws = wb.create_sheet(sheet_name) if sheet_name != "Aragua" else wb.active
        ws.title = sheet_name
        # Decorative title row above the header to exercise the header scan.
        ws.append(["REPORTE DE FLOTA - " + sheet_name.upper()])
        ws.append(HEADERS)
        ws.append(["04141234567", "GS-900 Wanwaytech", "ABC123", "", ""])
        ws.append(["04149876543", "VT08F - Protrack", "XYZ789", "", ""])

    wb.save(path)
    return wb


@pytest.fixture
def xlsx(tmp_path):
    path = tmp_path / "lote.xlsx"
    _make_workbook(path)
    return path


def test_load_devices_metadata_and_values(xlsx):
    df = load_devices(str(xlsx), commands_config=COMMANDS)
    assert len(df) == 4

    aragua = df[df["__sheet"] == "Aragua"].reset_index(drop=True)
    carabobo = df[df["__sheet"] == "Carabobo"].reset_index(drop=True)

    # Header is on row 2 (row 1 is decorative), so data starts at row 3.
    assert aragua.iloc[0]["__row"] == 3
    assert aragua.iloc[1]["__row"] == 4
    assert carabobo.iloc[0]["__row"] == 3
    assert carabobo.iloc[1]["__row"] == 4

    # Column indices: Telefono=1, Marca/Modelo=2, Placas=3, Status=4, Error=5.
    for _, row in df.iterrows():
        assert row["__status_col"] == 4
        assert row["__error_col"] == 5

    assert aragua.iloc[0]["Telefono"] == "04141234567"
    assert aragua.iloc[0]["Marca"] == "wanwaytech"
    assert aragua.iloc[0]["Modelo"] == "gs-900"
    assert aragua.iloc[0]["Placas"] == "ABC123"
    assert aragua.iloc[0]["Status"] == ""
    assert aragua.iloc[0]["Error"] == ""

    assert aragua.iloc[1]["Telefono"] == "04149876543"
    assert aragua.iloc[1]["Marca"] == "protrack"
    assert aragua.iloc[1]["Modelo"] == "vt08f"
    assert aragua.iloc[1]["Placas"] == "XYZ789"


def test_round_trip_preserves_values_and_formatting(xlsx):
    df = load_devices(str(xlsx), commands_config=COMMANDS)

    # Pre-set a style on the Status header and a data cell to verify it survives.
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["Aragua"]
    ws.cell(row=2, column=4).font = openpyxl.styles.Font(bold=True)
    ws.cell(row=3, column=2).fill = openpyxl.styles.PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    wb.save(xlsx)

    # Modify Status/Error in the DataFrame and save back.
    df.loc[df["__sheet"] == "Aragua", "Status"] = "ONLINE"
    df.loc[df["__sheet"] == "Aragua", "Error"] = ""
    df.loc[df["__sheet"] == "Carabobo", "Status"] = "OFFLINE"
    df.loc[df["__sheet"] == "Carabobo", "Error"] = "NO_RESPONSE_TIMEOUT"
    save_devices(df, str(xlsx))

    wb = openpyxl.load_workbook(xlsx)
    aragua = wb["Aragua"]
    carabobo = wb["Carabobo"]

    # Values written at the right rows/columns.
    assert aragua.cell(row=3, column=4).value == "ONLINE"
    assert aragua.cell(row=4, column=4).value == "ONLINE"
    assert carabobo.cell(row=3, column=4).value == "OFFLINE"
    assert carabobo.cell(row=3, column=5).value == "NO_RESPONSE_TIMEOUT"

    # Formatting preservation invariant.
    assert aragua.cell(row=2, column=4).font.bold is True
    assert aragua.cell(row=3, column=2).fill.start_color.rgb == "00FFFF00"


def test_parse_brand_model_vt08f_protrack():
    assert parse_brand_model("VT08F - Protrack", list(COMMANDS)) == ("protrack", "vt08f")


def test_parse_brand_model_gs900_wanwaytech():
    assert parse_brand_model("GS-900 Wanwaytech", list(COMMANDS)) == ("wanwaytech", "gs-900")


def test_parse_brand_model_empty():
    assert parse_brand_model("", list(COMMANDS)) == ("", "")
    assert parse_brand_model(None, list(COMMANDS)) == ("", "")


def test_unsupported_brand_still_loads_with_fallback(tmp_path):
    path = tmp_path / "lote.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aragua"
    ws.append(["REPORTE"])
    ws.append(HEADERS)
    ws.append(["04141234567", "Acme Widget", "ABC123", "", ""])
    wb.save(path)

    df = load_devices(str(path), commands_config=COMMANDS)
    assert len(df) == 1
    # Fallback: "modelo marca" -> brand = last token, model = remaining tokens.
    assert df.iloc[0]["Marca"] == "widget"
    assert df.iloc[0]["Modelo"] == "acme"
