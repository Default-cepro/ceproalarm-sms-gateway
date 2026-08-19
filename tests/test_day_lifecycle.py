from datetime import date

import openpyxl
import pytest

from src.services import day_lifecycle
from src.services.day_lifecycle import DeviceAggregate
from src.services.persistence import RunPersistence


class _Logger:
    def bind(self, **kwargs):
        return self

    def info(self, *a, **k):
        pass

    def warning(self, *a, **k):
        pass

    def error(self, *a, **k):
        pass

    def exception(self, *a, **k):
        pass

    def debug(self, *a, **k):
        pass

    def success(self, *a, **k):
        pass


class StubSMSService:
    def __init__(self, status="ONLINE", error_code=""):
        self.retries = 1
        self.delay = 0
        self.timeout = 1
        self.status = status
        self.error_code = error_code
        self.sent = []
        self.notifications = []

    async def send_with_retry(self, phone, message, expected):
        self.sent.append((phone, message, expected))
        return {"status": self.status, "error_code": self.error_code, "raw_message": "x"}

    async def send_notification(self, phone, message):
        self.notifications.append((phone, message))
        return {"status": "SENT", "message_id": "n1"}


class StubEmailReportService:
    def __init__(self):
        self.calls = []

    async def send_report(self, recipients, subject, body, attachment_paths):
        self.calls.append(
            {
                "recipients": recipients,
                "subject": subject,
                "body": body,
                "attachment_paths": attachment_paths,
            }
        )
        return {"sent_to": len(recipients), "attachments": len(attachment_paths), "subject": subject}


class StubPersistence:
    def __init__(self):
        self.cleared = False

    def clear(self):
        self.cleared = True


def _make_xlsx(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws.append(["Telefono", "Marca/Modelo", "Placas", "Status", "Error"])
    for row in rows:
        ws.append(row)
    wb.save(path)


def _build_states(path, logger):
    return day_lifecycle._prepare_daily_excel_states([str(path)], logger)


def test_normalize_status_valid():
    assert day_lifecycle._normalize_status("online") == "ONLINE"
    assert day_lifecycle._normalize_status("UNKNOWN") == "UNKNOWN"
    assert day_lifecycle._normalize_status("offline") == "OFFLINE"


def test_normalize_status_invalid_defaults_offline():
    assert day_lifecycle._normalize_status("") == "OFFLINE"
    assert day_lifecycle._normalize_status(None) == "OFFLINE"
    assert day_lifecycle._normalize_status("garbage") == "OFFLINE"


def test_merge_status_priority_offline_unknown_online():
    assert day_lifecycle._merge_status("OFFLINE", "UNKNOWN") == "UNKNOWN"
    assert day_lifecycle._merge_status("UNKNOWN", "ONLINE") == "ONLINE"
    assert day_lifecycle._merge_status("OFFLINE", "ONLINE") == "ONLINE"


def test_merge_status_keeps_higher():
    assert day_lifecycle._merge_status("ONLINE", "OFFLINE") == "ONLINE"
    assert day_lifecycle._merge_status("ONLINE", "UNKNOWN") == "ONLINE"
    assert day_lifecycle._merge_status("UNKNOWN", "OFFLINE") == "UNKNOWN"


def test_build_offline_alert_messages_empty():
    assert day_lifecycle._build_offline_alert_messages("2026-08-19", []) == []


def test_build_offline_alert_messages_content():
    messages = day_lifecycle._build_offline_alert_messages(
        "2026-08-19",
        [
            {
                "excel_path": "/tmp/lote.xlsx",
                "sheet": "Hoja1",
                "phone": "04122571528",
                "brand": "marca",
                "model": "modelo",
                "plate": "ABC123",
            }
        ],
    )
    assert len(messages) == 1
    msg = messages[0]
    assert "OFFLINE 2026-08-19" in msg
    assert "lote.xlsx" in msg
    assert "Hoja1" in msg
    assert "ABC123" in msg
    assert "MARCA" in msg
    assert "MODELO" in msg
    assert "04122571528" in msg


def test_build_offline_alert_messages_clipping():
    messages = day_lifecycle._build_offline_alert_messages(
        "2026-08-19",
        [
            {
                "excel_path": "/tmp/" + "x" * 100 + ".xlsx",
                "sheet": "s" * 100,
                "phone": "04122571528",
                "brand": "b" * 100,
                "model": "m" * 100,
                "plate": "p" * 100,
            }
        ],
        max_chars=60,
    )
    assert len(messages) == 1
    assert len(messages[0]) <= 60


def test_build_offline_alert_messages_sorted():
    messages = day_lifecycle._build_offline_alert_messages(
        "2026-08-19",
        [
            {"excel_path": "/tmp/b.xlsx", "sheet": "S", "phone": "2", "brand": "B", "model": "M", "plate": "Z"},
            {"excel_path": "/tmp/a.xlsx", "sheet": "S", "phone": "1", "brand": "A", "model": "M", "plate": "A"},
        ],
    )
    assert "a.xlsx" in messages[0]
    assert "b.xlsx" in messages[1]


async def test_execute_round_for_day_online(tmp_path):
    xlsx = tmp_path / "lote.xlsx"
    _make_xlsx(xlsx, [["04141234567", "protrack vt08f", "ABC123", "", ""]])
    logger = _Logger()
    states = _build_states(xlsx, logger)
    assert len(states) == 1
    state = states[0]
    idx = state.valid_indexes[0]

    sms = StubSMSService(status="ONLINE", error_code="")
    await day_lifecycle._execute_round_for_day(
        day_states=states,
        sms_service=sms,
        round_number=1,
        total_rounds=1,
        logger=logger,
    )

    assert len(sms.sent) == 1
    assert sms.sent[0][0] == "04141234567"
    assert state.aggregate[idx].status == "ONLINE"
    assert state.aggregate[idx].rounds_observed == 1
    assert state.aggregate[idx].error == ""


async def test_execute_round_online_then_offline_keeps_online(tmp_path):
    xlsx = tmp_path / "lote.xlsx"
    _make_xlsx(xlsx, [["04141234567", "protrack vt08f", "ABC123", "", ""]])
    logger = _Logger()
    states = _build_states(xlsx, logger)
    state = states[0]
    idx = state.valid_indexes[0]

    sms_online = StubSMSService(status="ONLINE", error_code="")
    await day_lifecycle._execute_round_for_day(
        day_states=states,
        sms_service=sms_online,
        round_number=1,
        total_rounds=2,
        logger=logger,
    )

    sms_offline = StubSMSService(status="OFFLINE", error_code="NO_RESPONSE_TIMEOUT")
    await day_lifecycle._execute_round_for_day(
        day_states=states,
        sms_service=sms_offline,
        round_number=2,
        total_rounds=2,
        logger=logger,
    )

    assert state.aggregate[idx].status == "ONLINE"
    assert state.aggregate[idx].rounds_observed == 2
    assert state.aggregate[idx].error == ""


async def test_resume_applies_round_result_for_first_index(tmp_path):
    xlsx = tmp_path / "lote.xlsx"
    _make_xlsx(
        xlsx,
        [
            ["04141234567", "protrack vt08f", "ABC123", "", ""],
            ["04149876543", "protrack vt08f", "XYZ789", "", ""],
        ],
    )
    logger = _Logger()
    persistence = RunPersistence(tmp_path / "run_state.json", logger)
    persistence.ensure_day("2026-08-19", ["08:00:00"], [str(xlsx)])
    # Simulate a mid-round kill: the first device (DataFrame index 0) was done.
    persistence.record_round_result(0, str(xlsx), 0, "ONLINE", "")

    states = day_lifecycle._prepare_daily_excel_states([str(xlsx)], logger, persistence=persistence)
    state = states[0]
    sms = StubSMSService(status="ONLINE", error_code="")
    await day_lifecycle._execute_round_for_day(
        day_states=states,
        sms_service=sms,
        round_number=1,
        total_rounds=1,
        logger=logger,
        persistence=persistence,
    )

    assert len(sms.sent) == 1
    assert state.aggregate[0].status == "ONLINE"
    assert state.aggregate[1].status == "ONLINE"


async def test_finalize_day_writes_status_and_notifies(tmp_path):
    xlsx = tmp_path / "lote.xlsx"
    _make_xlsx(xlsx, [["04141234567", "protrack vt08f", "ABC123", "", ""]])
    logger = _Logger()
    states = _build_states(xlsx, logger)
    state = states[0]
    idx = state.valid_indexes[0]
    state.aggregate[idx] = DeviceAggregate(status="OFFLINE", error="NO_RESPONSE_TIMEOUT", rounds_observed=1)

    sms = StubSMSService()
    email = StubEmailReportService()
    persistence = StubPersistence()

    await day_lifecycle._finalize_day(
        day_date=date(2026, 8, 19),
        day_states=states,
        sms_service=sms,
        offline_alert_recipients=["04143417356"],
        email_service=email,
        email_report_recipients=["ops@example.com"],
        email_subject_prefix="Ceproalarm",
        logger=logger,
        persistence=persistence,
    )

    # Status/Error written back to the xlsx.
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["Hoja1"]
    assert ws.cell(row=2, column=4).value == "OFFLINE"
    assert ws.cell(row=2, column=5).value == "NO_RESPONSE_TIMEOUT"

    # Offline alert sent to the recipient with the device details.
    assert len(sms.notifications) == 1
    recipient, message = sms.notifications[0]
    assert recipient == "04143417356"
    assert "ABC123" in message
    assert "04141234567" in message

    # Email report dispatched with the xlsx attached.
    assert len(email.calls) == 1
    call = email.calls[0]
    assert call["recipients"] == ["ops@example.com"]
    assert str(xlsx) in call["attachment_paths"]

    # Persistence cleared at day close.
    assert persistence.cleared is True
